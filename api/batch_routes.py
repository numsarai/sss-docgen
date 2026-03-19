"""
batch_routes.py — Excel batch upload and per-record document generation.

POST /batches/upload                          parse Excel → Batch + BatchRecords
GET  /batches                                 list all batches
GET  /batches/{id}                            batch + record summary
GET  /batches/{id}/records                    all records in batch
GET  /batches/{id}/export-excel               export all records as Excel (round-trip)
PUT  /batches/{id}/records/{rid}              update record variables → status=edited
POST /batches/{id}/records/{rid}/generate     generate DOCX+PDF for one record (sync)
POST /batches/{id}/generate-all               queue generation for all draft/edited records
DELETE /batches/{id}                          delete batch + all records

GET  /banks                                   Thai bank reference list
"""

import io
import json
import tempfile
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, BackgroundTasks, Depends, File, Form, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from .banks import THAI_BANKS
from .database import engine, get_db
from .generator import GENERATED_DIR, convert_to_pdf, generate_docx
from .models import (
    Batch, BatchListResponse, BatchRecord, BatchRecordResponse,
    BatchRecordUpdate, BatchResponse, CaseRecord,
)

router = APIRouter(tags=["batches"])

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _get_batch_or_404(db: Session, batch_id: int) -> Batch:
    b = db.get(Batch, batch_id)
    if b is None:
        raise HTTPException(status_code=404, detail=f"Batch {batch_id} not found.")
    return b


def _get_record_or_404(db: Session, batch_id: int, record_id: int) -> BatchRecord:
    r = db.get(BatchRecord, record_id)
    if r is None or r.batch_id != batch_id:
        raise HTTPException(status_code=404, detail=f"Record {record_id} not found in batch {batch_id}.")
    return r


def _parse_excel(data: bytes) -> list[dict]:
    """
    Parse an .xlsx file.  Row 1 = headers, rows 2+ = data.
    Returns a list of dicts {header: cell_value}.
    Empty rows (all cells blank) are skipped.
    """
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        return []

    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    records = []
    for row in rows[1:]:
        values = [str(v).strip() if v is not None else "" for v in row]
        if not any(values):          # skip blank rows
            continue
        records.append(dict(zip(headers, values)))
    return records


def _do_generate(batch_id: int, record_id: int) -> None:
    """Background-safe: generate DOCX+PDF for one BatchRecord and update DB."""
    from sqlalchemy.orm import Session as _Session
    with _Session(engine) as db:
        rec = db.get(BatchRecord, record_id)
        if rec is None or rec.batch_id != batch_id:
            return
        batch = db.get(Batch, batch_id)
        if batch is None:
            return

        try:
            # Create or reuse a CaseRecord
            if rec.case_id:
                case = db.get(CaseRecord, rec.case_id)
            else:
                case = None

            if case is None:
                case = CaseRecord(
                    template=batch.template,
                    _variables=json.dumps(rec.variables, ensure_ascii=False),
                    status="generating",
                )
                db.add(case)
                db.flush()
                rec.case_id = case.id

            case._variables = json.dumps(rec.variables, ensure_ascii=False)
            case.status = "generating"
            db.commit()

            docx_path = generate_docx(batch.template, rec.variables, case.id)
            pdf_path  = convert_to_pdf(docx_path)

            case.docx_path = str(docx_path)
            case.pdf_path  = str(pdf_path)
            case.status    = "generated"
            rec.status     = "generated"
            rec.case_id    = case.id
            db.commit()

        except Exception as exc:
            if rec.case_id:
                case = db.get(CaseRecord, rec.case_id)
                if case:
                    case.status = "error"
                    case.error  = str(exc)
            db.commit()


# ---------------------------------------------------------------------------
# GET /banks
# ---------------------------------------------------------------------------

@router.get("/banks", summary="List Thai bank reference data")
def list_banks():
    return THAI_BANKS


# ---------------------------------------------------------------------------
# POST /batches/upload
# ---------------------------------------------------------------------------

@router.post("/batches/upload", response_model=BatchResponse, status_code=status.HTTP_201_CREATED,
             summary="Upload Excel → create batch records")
def upload_batch(
    template: str        = Form(..., description="Path or name of the .docx template"),
    file:     UploadFile = File(..., description=".xlsx file to upload"),
    db:       Session    = Depends(get_db),
):
    if not (file.filename or "").endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=422, detail="Only .xlsx files are accepted.")

    data = file.file.read()
    try:
        rows = _parse_excel(data)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Failed to parse Excel: {exc}")

    if not rows:
        raise HTTPException(status_code=422, detail="Excel file is empty or has no data rows.")

    batch = Batch(filename=file.filename or "upload.xlsx", template=template)
    db.add(batch)
    db.flush()   # get batch.id before creating records

    for i, row in enumerate(rows, start=1):
        rec = BatchRecord(
            batch_id=batch.id,
            row_number=i,
            status="draft",
        )
        rec.variables = row
        db.add(rec)

    db.commit()
    db.refresh(batch)
    return BatchResponse.from_batch(batch)


# ---------------------------------------------------------------------------
# GET /batches
# ---------------------------------------------------------------------------

@router.get("/batches", response_model=BatchListResponse, summary="List all batches")
def list_batches(db: Session = Depends(get_db)):
    batches = db.query(Batch).order_by(Batch.created_at.desc()).all()
    return BatchListResponse(
        total=len(batches),
        items=[BatchResponse.from_batch(b) for b in batches],
    )


# ---------------------------------------------------------------------------
# GET /batches/{id}
# ---------------------------------------------------------------------------

@router.get("/batches/{batch_id}", response_model=BatchResponse, summary="Get batch summary")
def get_batch(batch_id: int, db: Session = Depends(get_db)):
    return BatchResponse.from_batch(_get_batch_or_404(db, batch_id))


# ---------------------------------------------------------------------------
# GET /batches/{id}/records
# ---------------------------------------------------------------------------

@router.get("/batches/{batch_id}/records", response_model=list[BatchRecordResponse],
            summary="List all records in a batch")
def list_records(batch_id: int, db: Session = Depends(get_db)):
    _get_batch_or_404(db, batch_id)
    records = (
        db.query(BatchRecord)
        .filter(BatchRecord.batch_id == batch_id)
        .order_by(BatchRecord.row_number)
        .all()
    )
    return [BatchRecordResponse.from_record(r) for r in records]


# ---------------------------------------------------------------------------
# GET /batches/{id}/export-excel  — round-trip Excel export
# ---------------------------------------------------------------------------

@router.get("/batches/{batch_id}/export-excel", summary="Export all records as Excel")
def export_batch_excel(batch_id: int, db: Session = Depends(get_db)):
    """
    Export every record in the batch as an Excel file.

    Columns:
      record_id | row_number | status | <all variable columns in original order>

    All edits are reflected; Thai text is preserved natively in xlsx UTF-8.
    """
    batch = _get_batch_or_404(db, batch_id)
    records = (
        db.query(BatchRecord)
        .filter(BatchRecord.batch_id == batch_id)
        .order_by(BatchRecord.row_number)
        .all()
    )
    if not records:
        raise HTTPException(status_code=422, detail="Batch has no records to export.")

    # Collect variable columns in original insertion order (first record defines order,
    # later records may introduce extra keys which are appended).
    col_order: list[str] = []
    seen: set[str] = set()
    for rec in records:
        for key in rec.variables.keys():
            if key not in seen:
                col_order.append(key)
                seen.add(key)

    meta_cols   = ["record_id", "row_number", "status"]
    all_headers = meta_cols + col_order

    # ---- build workbook -----------------------------------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Records"

    header_fill  = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    meta_fill    = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
    draft_fill   = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    edited_fill  = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    gen_fill     = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")

    STATUS_FILL = {"draft": draft_fill, "edited": edited_fill, "generated": gen_fill}

    # Header row
    for col_idx, header in enumerate(all_headers, start=1):
        is_meta = col_idx <= len(meta_cols)
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = Font(bold=True, color="FFFFFF", name="TH Sarabun New", size=11)
        cell.fill      = meta_fill if is_meta else header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    for row_idx, rec in enumerate(records, start=2):
        row_fill = STATUS_FILL.get(rec.status, draft_fill)

        ws.cell(row=row_idx, column=1, value=rec.id).fill          = row_fill
        ws.cell(row=row_idx, column=2, value=rec.row_number).fill  = row_fill
        ws.cell(row=row_idx, column=3, value=rec.status).fill      = row_fill

        for col_idx, col_name in enumerate(col_order, start=len(meta_cols) + 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=rec.variables.get(col_name, ""))
            cell.fill      = row_fill
            cell.font      = Font(name="TH Sarabun New", size=11)
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Column widths
    for col_idx, header in enumerate(all_headers, start=1):
        col_letter = get_column_letter(col_idx)
        width = max(len(header) * 1.8, 12)
        ws.column_dimensions[col_letter].width = min(width, 35)

    ws.freeze_panes      = "A2"
    ws.row_dimensions[1].height = 26

    # ---- stream -------------------------------------------------------
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote as _quote
    safe_name = batch.filename.rsplit(".", 1)[0].replace(" ", "_")
    filename  = f"{safe_name}_export.xlsx"
    encoded   = _quote(filename, safe="")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


# ---------------------------------------------------------------------------
# PUT /batches/{id}/records/{rid}
# ---------------------------------------------------------------------------

@router.put("/batches/{batch_id}/records/{record_id}", response_model=BatchRecordResponse,
            summary="Update record variables")
def update_record(
    batch_id:  int,
    record_id: int,
    body:      BatchRecordUpdate,
    db:        Session = Depends(get_db),
):
    rec = _get_record_or_404(db, batch_id, record_id)
    rec.variables = body.variables
    if rec.status == "draft":
        rec.status = "edited"
    db.commit()
    db.refresh(rec)
    return BatchRecordResponse.from_record(rec)


# ---------------------------------------------------------------------------
# POST /batches/{id}/records/{rid}/generate
# ---------------------------------------------------------------------------

@router.post("/batches/{batch_id}/records/{record_id}/generate",
             response_model=BatchRecordResponse,
             summary="Generate DOCX+PDF for a single record (synchronous)")
def generate_record(
    batch_id:  int,
    record_id: int,
    db:        Session = Depends(get_db),
):
    rec = _get_record_or_404(db, batch_id, record_id)
    # Run synchronously — LibreOffice takes a few seconds
    _do_generate(batch_id, record_id)
    db.expire(rec)
    db.refresh(rec)
    return BatchRecordResponse.from_record(rec)


# ---------------------------------------------------------------------------
# POST /batches/{id}/generate-all
# ---------------------------------------------------------------------------

@router.post("/batches/{batch_id}/generate-all",
             summary="Queue generation for all draft/edited records")
def generate_all(
    batch_id:        int,
    background_tasks: BackgroundTasks,
    db:              Session = Depends(get_db),
):
    _get_batch_or_404(db, batch_id)
    pending = (
        db.query(BatchRecord)
        .filter(BatchRecord.batch_id == batch_id, BatchRecord.status.in_(["draft", "edited"]))
        .order_by(BatchRecord.row_number)
        .all()
    )
    if not pending:
        return {"queued": 0, "message": "No draft or edited records to generate."}

    for rec in pending:
        background_tasks.add_task(_do_generate, batch_id, rec.id)

    return {"queued": len(pending), "message": f"Queued {len(pending)} records for generation."}


# ---------------------------------------------------------------------------
# POST /batches/{id}/records/{rid}/discard  — reset generated record back to draft
# ---------------------------------------------------------------------------

@router.post("/batches/{batch_id}/records/{record_id}/discard",
             response_model=BatchRecordResponse,
             summary="Discard generated preview — deletes associated case, resets to draft")
def discard_record(
    batch_id:  int,
    record_id: int,
    db:        Session = Depends(get_db),
):
    """
    Remove the generated preview for a single batch record.
    - Deletes the associated CaseRecord (docx + pdf files stay on disk for now)
    - Resets BatchRecord.status back to "draft" (or "edited" if it was edited)
    - Clears case_id
    """
    rec = _get_record_or_404(db, batch_id, record_id)

    if rec.case_id:
        case = db.get(CaseRecord, rec.case_id)
        if case:
            db.delete(case)

    # Reset: go back to "edited" if variables were modified, else "draft"
    rec.status  = "edited" if rec.status in ("generated",) and rec.case_id else "draft"
    rec.case_id = None
    rec.status  = "draft"     # always full reset so user can re-generate
    db.commit()
    db.refresh(rec)
    return BatchRecordResponse.from_record(rec)


# ---------------------------------------------------------------------------
# DELETE /batches/{id}
# ---------------------------------------------------------------------------

@router.delete("/batches/{batch_id}", status_code=status.HTTP_204_NO_CONTENT,
               summary="Delete batch and all its records")
def delete_batch(batch_id: int, db: Session = Depends(get_db)):
    batch = _get_batch_or_404(db, batch_id)
    db.delete(batch)
    db.commit()
