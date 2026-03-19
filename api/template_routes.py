"""
template_routes.py — CRUD endpoints for .docx template management.

GET    /templates                          list all templates (paginated)
POST   /templates                          upload a new .docx template
GET    /templates/{id}                     get a single template
PUT    /templates/{id}                     edit name/description, optionally re-upload file
DELETE /templates/{id}                     delete template record + file
GET    /templates/{id}/download            download the .docx file
GET    /templates/{id}/excel-template      download an Excel template pre-filled with column headers
"""

import io
import shutil
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session

from .database import get_db
from .extractor import scan_template_fields
from .models import TemplateList, TemplateRecord, TemplateResponse

# Templates are stored here (alongside the project)
_TEMPLATES_DIR = Path(__file__).parent.parent / "templates"
_TEMPLATES_DIR.mkdir(exist_ok=True)

router = APIRouter(prefix="/templates", tags=["templates"])


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _get_or_404(db: Session, template_id: int) -> TemplateRecord:
    record = db.get(TemplateRecord, template_id)
    if record is None:
        raise HTTPException(status_code=404, detail=f"Template {template_id} not found.")
    return record


def _resolve_template_path(stored_path: str) -> Optional[Path]:
    """
    Return an existing Path for a template file.

    Falls back to looking up by filename inside _TEMPLATES_DIR so that
    records created before a Docker migration (with host-absolute paths)
    still resolve correctly inside the container.
    """
    p = Path(stored_path)
    if p.exists():
        return p
    candidate = _TEMPLATES_DIR / p.name
    if candidate.exists():
        return candidate
    return None


def _save_file(upload: UploadFile, existing_path: Optional[str] = None) -> Path:
    """Save uploaded file to the templates directory, removing any old file."""
    if existing_path:
        old = Path(existing_path)
        if old.exists():
            old.unlink(missing_ok=True)

    safe_name = Path(upload.filename or "template.docx").name
    dest = _TEMPLATES_DIR / safe_name

    # Avoid collisions
    counter = 1
    stem, suffix = dest.stem, dest.suffix
    while dest.exists():
        dest = _TEMPLATES_DIR / f"{stem}_{counter}{suffix}"
        counter += 1

    with dest.open("wb") as f:
        shutil.copyfileobj(upload.file, f)
    return dest


# ---------------------------------------------------------------------------
# GET /templates
# ---------------------------------------------------------------------------

@router.get("", response_model=TemplateList)
def list_templates(
    skip:  int = Query(default=0,  ge=0),
    limit: int = Query(default=50, ge=1, le=200),
    db:    Session = Depends(get_db),
):
    total = db.query(func.count(TemplateRecord.id)).scalar()
    records = (
        db.query(TemplateRecord)
        .order_by(TemplateRecord.created_at.desc())
        .offset(skip)
        .limit(limit)
        .all()
    )
    return TemplateList(
        total=total, skip=skip, limit=limit,
        items=[TemplateResponse.from_record(r) for r in records],
    )


# ---------------------------------------------------------------------------
# POST /templates
# ---------------------------------------------------------------------------

@router.post("", response_model=TemplateResponse, status_code=status.HTTP_201_CREATED)
def create_template(
    name:        str        = Form(...),
    description: str        = Form(default=""),
    file:        UploadFile = File(...),
    db:          Session    = Depends(get_db),
):
    if not file.filename or not file.filename.endswith(".docx"):
        raise HTTPException(status_code=422, detail="Only .docx files are accepted.")

    dest = _save_file(file)

    try:
        fields = scan_template_fields(dest)
    except Exception:
        fields = []

    record = TemplateRecord(
        name=name,
        description=description or None,
        path=str(dest),
    )
    record.fields = fields
    db.add(record)
    db.commit()
    db.refresh(record)
    return TemplateResponse.from_record(record)


# ---------------------------------------------------------------------------
# GET /templates/{id}
# ---------------------------------------------------------------------------

@router.get("/{template_id}", response_model=TemplateResponse)
def get_template(template_id: int, db: Session = Depends(get_db)):
    return TemplateResponse.from_record(_get_or_404(db, template_id))


# ---------------------------------------------------------------------------
# PUT /templates/{id}
# ---------------------------------------------------------------------------

@router.put("/{template_id}", response_model=TemplateResponse)
def update_template(
    template_id: int,
    name:        str                  = Form(...),
    description: str                  = Form(default=""),
    file:        Optional[UploadFile] = File(default=None),
    db:          Session              = Depends(get_db),
):
    record = _get_or_404(db, template_id)

    record.name        = name
    record.description = description or None

    if file and file.filename:
        if not file.filename.endswith(".docx"):
            raise HTTPException(status_code=422, detail="Only .docx files are accepted.")
        dest = _save_file(file, existing_path=record.path)
        record.path = str(dest)
        try:
            record.fields = scan_template_fields(dest)
        except Exception:
            pass

    db.commit()
    db.refresh(record)
    return TemplateResponse.from_record(record)


# ---------------------------------------------------------------------------
# DELETE /templates/{id}
# ---------------------------------------------------------------------------

@router.delete("/{template_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_template(template_id: int, db: Session = Depends(get_db)):
    record = _get_or_404(db, template_id)
    path = Path(record.path)
    db.delete(record)
    db.commit()
    if path.exists():
        path.unlink(missing_ok=True)


# ---------------------------------------------------------------------------
# GET /templates/{id}/download
# ---------------------------------------------------------------------------

@router.get("/{template_id}/download")
@router.get("/{template_id}/download/{token_path}")
def download_template(template_id: int, db: Session = Depends(get_db), token_path: str = ''):
    record = _get_or_404(db, template_id)
    tpl_path = _resolve_template_path(record.path)
    if tpl_path is None:
        raise HTTPException(status_code=404, detail="Template file not found on disk.")
    record.path = str(tpl_path)
    db.commit()
    return FileResponse(
        str(tpl_path),
        filename=tpl_path.name,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )


# ---------------------------------------------------------------------------
# GET /templates/{id}/excel-template  — generate an Excel fill-in template
# ---------------------------------------------------------------------------

# Known Thai labels for common variable names used in Thai police templates.
_THAI_LABELS: dict[str, str] = {
    "เลขที่หนังสือ":           "เลขที่หนังสือ",
    "วันที่":                  "วันที่",
    "เดือน":                   "เดือน",
    "ปีพ.ศ.":                  "ปี พ.ศ.",
    "เลขบัญชี":                "เลขบัญชี",
    "ชื่อบัญชี":               "ชื่อบัญชี",
    "ชื่อผู้เสียหาย":          "ชื่อผู้เสียหาย",
    "วันที่แจ้งความ":          "วันที่แจ้งความ",
    "จำนวนครั้งโอน":           "จำนวนครั้งโอน",
    "จำนวนเงินหลัก":           "จำนวนเงิน (บาท)",
    "จำนวนเงินสตางค์":         "จำนวนเงิน (สตางค์)",
    "BANK_ID":                 "รหัสธนาคาร (BANK_ID)",
    "รหัสธนาคาร":             "รหัสธนาคาร",
    "รหัสสาขา":                "รหัสสาขา",
    "CASE_ID":                 "รหัสคดี (CASE_ID)",
    "วันเริ่มต้นStatement":    "วันเริ่มต้น Statement",
    "วันสิ้นสุดStatement":     "วันสิ้นสุด Statement",
    "วันรับโอนเงิน":           "วันรับโอนเงิน",
    "ยศพนักงานสอบสวน":        "ยศพนักงานสอบสวน",
    "ชื่อพนักงานสอบสวน":      "ชื่อพนักงานสอบสวน",
    "ตำแหน่งพนักงานสอบสวน":  "ตำแหน่งพนักงานสอบสวน",
    "โทรศัพท์":                "โทรศัพท์",
    "อีเมลพนักงานสอบสวน":     "อีเมล",
    "วันกำหนดส่งเอกสาร":      "วันกำหนดส่งเอกสาร",
    "เวลากำหนดส่ง":            "เวลากำหนดส่ง",
    "ยศผู้ลงนาม":              "ยศผู้ลงนาม",
    "ชื่อผู้ลงนาม":            "ชื่อผู้ลงนาม",
}


@router.get("/{template_id}/excel-template")
def download_excel_template(template_id: int, db: Session = Depends(get_db)):
    """
    Generate and return an Excel (.xlsx) fill-in template whose columns match
    the {{variables}} found in the .docx template.

    Row 1 — field names (variable names as-is)
    Row 2 — Thai display labels (where known, otherwise same as row 1)
    Row 3 — empty example row for the user to fill in
    """
    record = _get_or_404(db, template_id)
    fields = record.fields or []
    if not fields:
        raise HTTPException(status_code=422, detail="Template has no fields to export.")

    # ---- build workbook -----------------------------------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    header_fill  = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    label_fill   = PatternFill(start_color="2E5C8A", end_color="2E5C8A", fill_type="solid")
    example_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")

    for col_idx, field in enumerate(fields, start=1):
        col_letter = get_column_letter(col_idx)

        # Row 1 — machine name (bold white on dark blue)
        cell1 = ws.cell(row=1, column=col_idx, value=field)
        cell1.font      = Font(bold=True, color="FFFFFF", name="TH Sarabun New", size=11)
        cell1.fill      = header_fill
        cell1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Row 2 — Thai label (white on medium blue)
        label = _THAI_LABELS.get(field, field)
        cell2 = ws.cell(row=2, column=col_idx, value=label)
        cell2.font      = Font(color="FFFFFF", name="TH Sarabun New", size=10)
        cell2.fill      = label_fill
        cell2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Row 3 — example placeholder (light blue bg, user fills here)
        cell3 = ws.cell(row=3, column=col_idx, value="")
        cell3.fill      = example_fill
        cell3.font      = Font(name="TH Sarabun New", size=11)
        cell3.alignment = Alignment(horizontal="left", vertical="center")

        # Column width: generous enough for Thai characters
        width = max(len(field) * 2, len(label) * 1.5, 14)
        ws.column_dimensions[col_letter].width = min(width, 30)

    # Freeze first two header rows
    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22

    # ---- stream response ----------------------------------------------
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote as _quote
    safe_name = record.name.replace(" ", "_").replace("/", "-")
    filename  = f"{safe_name}_template.xlsx"
    encoded   = _quote(filename, safe="")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


# ---------------------------------------------------------------------------
# GET /templates/{id}/editor-config  — OnlyOffice editor configuration
# ---------------------------------------------------------------------------

@router.get("/{template_id}/editor-config")
def template_editor_config(template_id: int, db: Session = Depends(get_db)):
    import os
    import time
    from docx import Document as DocxDoc
    from .generator import _apply_thai_numerals

    record = _get_or_404(db, template_id)
    tpl_path = _resolve_template_path(record.path)
    if tpl_path is None:
        raise HTTPException(status_code=404, detail="Template file not found on disk.")
    record.path = str(tpl_path)
    db.commit()

    # Convert any Arabic digits → Thai before opening in editor
    doc = DocxDoc(str(tpl_path))
    _apply_thai_numerals(doc)
    doc.save(str(tpl_path))

    key = f"template-{template_id}-{int(time.time())}"
    token = os.getenv("APP_PASSWORD", "secret")
    return {
        "document": {
            "fileType": "docx",
            "key": key,
            "title": f"{record.name}.docx",
            "url": f"http://host.docker.internal:8000/templates/{template_id}/download/{token}",
        },
        "documentType": "word",
        "editorConfig": {
            "callbackUrl": f"http://host.docker.internal:8000/templates/{template_id}/editor-callback/{token}",
            "mode": "edit",
            "lang": "th",
            "customization": {"autosave": True, "forcesave": True},
        },
    }


# ---------------------------------------------------------------------------
# POST /templates/{id}/editor-callback  — OnlyOffice saves the document here
# ---------------------------------------------------------------------------

@router.post("/{template_id}/editor-callback")
@router.post("/{template_id}/editor-callback/{token_path}")
def template_editor_callback(template_id: int, body: dict, db: Session = Depends(get_db), token_path: str = ''):
    """
    OnlyOffice calls this URL when the user saves.
    status 2 = ready to save, status 6 = force-saved.
    Must return {"error": 0} on success.
    """
    import urllib.request as urlreq

    if body.get("status") in (2, 6):
        url = body.get("url")
        if not url:
            return {"error": 1}

        record = _get_or_404(db, template_id)

        try:
            with urlreq.urlopen(url) as resp:
                data = resp.read()

            Path(record.path).write_bytes(data)

            # Rescan fields since content may have changed
            try:
                record.fields = scan_template_fields(record.path)
            except Exception:
                pass

            db.commit()
        except Exception:
            return {"error": 1}

    return {"error": 0}
